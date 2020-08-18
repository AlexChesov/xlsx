#! python3
# -*- coding: utf-8 -*-
import os
import time
import sys
import pandas as pd
import openpyxl, pprint

print('Открытие рабочей книги...')
start = time.time()
wb = openpyxl.load_workbook('/Users/alexchesov/py/xlsx/deficitka_value.xlsx')# ,data_only=True
endtime = (time.time() - start)/60
print('Время открытия книги = ' + str(endtime) + ' мин')
sheets = wb.sheetnames
print('Определение активного листа')
wb.active = 1
sheet = wb.active
print('Активный лист - ' + str(sheet))



global pp_size# Объем ПП шт.
global pp_demand# Производственная программа
global pp_demandTD
global pp_demandHM
global pp_sizeAJR# Объем ПП шт.
global pp_demandAJR# Производственная программа
global pp_demand_sumAJR# Потребность производственной программы
global pp_demand_sumTD
global pp_demand_sumHM
global currency_rub
global currency_usd
global currency_eur
global currency_gbp
global currency_cny



sheet_currency = wb['Валюта']
currency_rub = sheet_currency['B2'].value
currency_usd = sheet_currency['B3'].value
currency_eur = sheet_currency['B4'].value
currency_gbp = sheet_currency['B5'].value
currency_cny = sheet_currency['B6'].value

start = time.time()

for i in range (15, sheet.max_row+1):
	pp_demand_sum = 0
	pp_demand_sumTD = 0
	pp_demand = 0
	pp_demandTD = 0
	cur_cur = 0
	cur_detect = sheet.cell(row = i, column = 8)
	if str(cur_detect.value) in ['Руб','руб','Руб.','руб.']:
		cur_cur = currency_rub
	elif str(cur_detect.value) in ['EUR','eur']:
		cur_cur = currency_eur
	elif cur_detect.value in ['USD','usd']:
		cur_cur = currency_usd
	elif cur_detect.value in ['GBP','gbp']:
		cur_cur = currency_gbp
	elif cur_detect.value in ['CNY','cny']:
		cur_cur = currency_cny
	vom = sheet.cell(row = i, column = 15)
	# print(cur_cur)
	price_cur = sheet.cell(row = i, column = 9)
	set_cur = sheet.cell(row = i, column = 14)
	try:
		set_cur.value = cur_cur
		vom.value = cur_cur * price_cur.value
	except TypeError:
		pass
	for j in range (19, 118+1):# Вычисление Потребности производственной программы, колонка DP и колонка TD и колонка HM
		pp_size = sheet.cell(row = 10, column = j)
		try:
			if int(pp_size.value) > 0:
				machina_num = sheet.cell(row = i, column = j)
				pp_demand = float(pp_size.value) * float(machina_num.value)
				pp_demandTD = float(pp_size.value) * float(machina_num.value) * float(vom.value)
				pp_demand_sum += pp_demand
				pp_demand_sumTD += pp_demandTD
				cell = sheet.cell(row = i, column = j+102)
				cellTD = sheet.cell(row = i, column = j + 506)
				cell.value = pp_demand
				cellTD.value = pp_demandTD
		except TypeError:
			pass
		# except ValueError:
		# 	pass
	cell_sum = sheet.cell(row = i, column = 120)
	cellTD_sum = sheet.cell(row = i, column = 524)
	cell_sum.value = pp_demand_sum
	cellTD_sum.value = pp_demand_sumTD
	pp_demand_sum = 0
	pp_demand = 0
	pp_demand_sumTD = 0
	pp_demandTD = 0
	for j in range(854, 953+1):# Вычисление Потребности производственной программы, колонка AJR
		pp_size = sheet.cell(row = 10, column = j)
		try:
			if int(pp_size.value) > 0:
				machina_num = sheet.cell(row = i, column = j)
				pp_demand = float(pp_size.value) * float(machina_num.value)
				pp_demand_sum += pp_demand
				cell = sheet.cell(row = i, column = j+101)
				cell.value = pp_demand
		except TypeError:
			pass
	cell_sum = sheet.cell(row = i, column = 954)
	cell_sum.value = pp_demand_sum
	pp_demand = 0
	pp_demand_sum = 0
endtime = time.time() - start
print('Время вычислений и записи результатов = ' + str(endtime) + ' сек')
print('Сохраняем книгу...')
start = time.time()
wb.save('/Users/alexchesov/py/xlsx/deficitka_value_full_sum_v2.xlsx')
endtime = (time.time() - start)/60
print('Время сохранения книги = ' + str(endtime) + ' мин')